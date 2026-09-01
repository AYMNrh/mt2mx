#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
from pathlib import Path
import subprocess
import sys
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
EXPECTED_MT = {"MT103": 23, "MT202": 11, "MT202_COV": 19, "MT910": 9, "MT920": 5}


def read_csv(name: str) -> list[dict[str, str]]:
    with (ROOT / name).open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def main() -> None:
    failures: list[str] = []

    def check(condition: bool, message: str) -> None:
        if not condition:
            failures.append(message)

    report = json.loads((ROOT / "outputs/completeness_report.json").read_text(encoding="utf-8"))
    mt = read_csv("outputs/mt_field_inventory.csv")
    targets = read_csv("outputs/mx_target_inventory.csv")
    crosswalk = read_csv("outputs/mt_to_mx_crosswalk.csv")
    lineage = read_csv("outputs/mt_to_mx_to_dfr_lineage.csv")
    queue = read_csv("outputs/review_queue.csv")
    issues = read_csv("outputs/dq_issue_log.csv")
    sources = read_csv("sources/source_manifest.csv")

    actual_mt = {message: sum(row["message_type"] == message for row in mt) for message in EXPECTED_MT}
    check(actual_mt == EXPECTED_MT, f"MT count mismatch: {actual_mt}")
    check(len(mt) == 67, f"expected 67 MT occurrences, got {len(mt)}")
    source_keys = {(row["message_type"], int(row["occurrence_no"])) for row in mt}
    crosswalk_keys = {(row["mt_message"], int(row["mt_occurrence_no"])) for row in crosswalk}
    check(source_keys == crosswalk_keys, "crosswalk source coverage does not equal MT inventory")
    check(len(queue) == 67, f"expected 67 review rows, got {len(queue)}")
    check(len(crosswalk) == 234, f"expected 234 crosswalk rows, got {len(crosswalk)}")
    check(sum(bool(row["mx_path"]) for row in crosswalk) == 233, "expected 233 non-empty MX target rows")

    target_keys = {(row["mx_message_id"], row["path"]) for row in targets}
    invalid_targets = [
        (row["mt_message"], row["mt_occurrence_no"], row["mx_message_id"], row["mx_path"])
        for row in crosswalk
        if row["mx_path"] and (row["mx_message_id"], row["mx_path"]) not in target_keys
    ]
    check(not invalid_targets, f"crosswalk has unpublished/invalid target paths: {invalid_targets[:5]}")

    status_counts: dict[str, int] = {}
    for row in lineage:
        status_counts[row["dfr_join_status"]] = status_counts.get(row["dfr_join_status"], 0) + 1
    expected_status = {"MATCHED": 222, "DFR_MESSAGE_NOT_IN_SCOPE": 10, "PATH_NOT_MAPPED": 1, "NOT_APPLICABLE": 1}
    check(status_counts == expected_status, f"DFR status mismatch: {status_counts}")
    check(report["dfr_status_counts"] == expected_status, "JSON report DFR counts disagree with lineage")
    check(report["verification"]["all_67_mt_occurrences_covered"] is True, "report coverage flag is false")
    check(report["verification"]["cbpr_approved"] is False, "report must not claim CBPR approval")

    blocking_ids = {row["source_id"] for row in sources if row["review_status"] == "BLOCKING"}
    check(blocking_ids == {"CBPR_USAGE_GUIDELINES", "SWIFT_TRANSLATION_LIBRARY"}, f"blocking source set mismatch: {blocking_ids}")
    local_hash_rows = [row for row in sources if row["source_kind"] in {"PDF", "XSD_STRUCTURAL_CORROBORATION", "INTERNAL_DERIVED_MAPPING"}]
    bad_hashes = [row["source_id"] for row in local_hash_rows if len(row["sha256"]) != 64]
    check(not bad_hashes, f"missing/invalid source hashes: {bad_hashes}")
    check(len(issues) == 10, f"expected 10 DQ issues, got {len(issues)}")

    html = (ROOT / "docs/index.html").read_text(encoding="utf-8")
    check("MT → MX → DFR review pack" in html, "HTML title/content missing")
    check("Not production-approved" in html, "HTML approval warning missing")

    try:
        from openpyxl import load_workbook
        workbook = load_workbook(ROOT / "outputs/MT_TO_MX_REVIEW.xlsx", read_only=True, data_only=False)
        expected_sheets = {"Executive Summary", "Source Register", "MT Field Inventory", "MT to MX Crosswalk", "MX to DFR Lineage", "Review Queue", "DQ Issues", "MX Target Inventory"}
        check(set(workbook.sheetnames) == expected_sheets, f"workbook sheets mismatch: {workbook.sheetnames}")
        formulas = [
            f"{ws.title}!{cell.coordinate}"
            for ws in workbook.worksheets for row in ws.iter_rows() for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        ]
        check(not formulas, f"unexpected workbook formulas: {formulas[:5]}")
        workbook.close()
    except Exception as exc:
        failures.append(f"workbook verification failed: {exc}")

    tracked = subprocess.check_output(["git", "ls-files"], cwd=ROOT, text=True).splitlines()
    unsafe = [
        name for name in tracked
        if name.startswith(".private/") or Path(name).suffix.lower() in {".pdf", ".txt", ".xsd"}
    ]
    check(not unsafe, f"private/licensed file types are tracked: {unsafe}")

    required_files = [
        "README.md",
        "docs/index.html",
        "docs/MT_TO_MX_TO_DFR_GUIDE.md",
        "sources/source_manifest.csv",
        "outputs/mt_field_inventory.csv",
        "outputs/mx_target_inventory.csv",
        "outputs/mt_to_mx_crosswalk.csv",
        "outputs/mt_to_mx_to_dfr_lineage.csv",
        "outputs/review_queue.csv",
        "outputs/dq_issue_log.csv",
        "outputs/MT_TO_MX_REVIEW.xlsx",
        "outputs/completeness_report.json",
    ]
    missing = [name for name in required_files if not (ROOT / name).exists()]
    check(not missing, f"missing required artifacts: {missing}")

    result: dict[str, Any] = {
        "status": "PASS" if not failures else "FAIL",
        "checks": {
            "mt_occurrences": len(mt),
            "crosswalk_rows": len(crosswalk),
            "selected_mx_paths": len(targets),
            "lineage_rows": len(lineage),
            "dfr_status_counts": status_counts,
            "source_rows": len(sources),
            "dq_issues": len(issues),
            "unsafe_tracked_files": unsafe,
        },
        "failures": failures,
    }
    print(json.dumps(result, indent=2, sort_keys=True))
    raise SystemExit(1 if failures else 0)


if __name__ == "__main__":
    main()
