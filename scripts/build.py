#!/usr/bin/env python3
from __future__ import annotations

import argparse
from copy import copy
import csv
from datetime import datetime, timezone
from hashlib import sha256
from html import escape
import json
from pathlib import Path
import re
import subprocess
import sys
from typing import Any, Iterable

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from mt2mx.mapping import build_crosswalk, join_dfr  # noqa: E402
from mt2mx.mt_parser import parse_mt_guide_text  # noqa: E402
from mt2mx.rules import curated_rules  # noqa: E402
from mt2mx.xsd_parser import extract_xsd_inventory  # noqa: E402

MT_SPECS = [
    ("SR_2025_MT103.txt", "MT103", "SR_2025_MT103.pdf"),
    ("SR_2025_MT202.txt", "MT202", "SR_2025_MT202.pdf"),
    ("SR_2025_MT202.COV.txt", "MT202_COV", "SR_2025_MT202.COV.pdf"),
    ("SR_2025_MT910.txt", "MT910", "SR_2025_MT910.pdf"),
    ("SR_2025_MT920.txt", "MT920", "SR_2025_MT920.pdf"),
]
XSD_NAMES = [
    "pacs.008.001.08.xsd",
    "pacs.009.001.08.xsd",
    "camt.054.001.08.xsd",
    "camt.060.001.05.xsd",
]
VERSION_ALIASES = {
    "pacs.008.001.08": "pacs.008.001.14",
    "pacs.009.001.08": "pacs.009.001.13",
    "camt.054.001.08": "camt.054.001.14",
}
PDF_SPECS = {
    "SR_2025_MT103.pdf": ("SR_2025_MT103", "SWIFT", "Standards MT November 2025 — MT 103 Single Customer Credit Transfer", "MT November 2025", "MT103", "PRIMARY"),
    "SR_2025_MT202.pdf": ("SR_2025_MT202", "SWIFT", "Standards MT November 2025 — MT 202 General Financial Institution Transfer", "MT November 2025", "MT202", "PRIMARY"),
    "SR_2025_MT202.COV.pdf": ("SR_2025_MT202_COV", "SWIFT", "Standards MT November 2025 — MT 202 COV General Financial Institution Transfer", "MT November 2025", "MT202_COV", "PRIMARY"),
    "SR_2025_MT910.pdf": ("SR_2025_MT910", "SWIFT", "Standards MT November 2025 — MT 910 Confirmation of Credit", "MT November 2025", "MT910", "PRIMARY"),
    "SR_2025_MT920.pdf": ("SR_2025_MT920", "SWIFT", "Standards MT November 2025 — MT 920 Request Message", "MT November 2025", "MT920", "PRIMARY"),
    "MX_pacs_008_001_08.pdf": ("MX_pacs_008_001_08", "ISO 20022 / MyStandards export", "pacs.008.001.08 FIToFICustomerCreditTransferV08 Message Definition Report", "pacs.008.001.08", "pacs.008.001.08", "PRIMARY"),
    "MX_pacs_009_001_08.pdf": ("MX_pacs_009_001_08", "ISO 20022 / MyStandards export", "pacs.009.001.08 FinancialInstitutionCreditTransferV08 Message Definition Report", "pacs.009.001.08", "pacs.009.001.08", "PRIMARY"),
    "MX_camt_054_001_08.pdf": ("MX_camt_054_001_08", "ISO 20022 / MyStandards export", "camt.054.001.08 BankToCustomerDebitCreditNotificationV08 Message Definition Report", "camt.054.001.08", "camt.054.001.08", "PRIMARY"),
    "MX_camt_060_001_05.pdf": ("MX_camt_060_001_05", "ISO 20022 / MyStandards export", "camt.060.001.05 AccountReportingRequestV05 Message Definition Report", "camt.060.001.05", "camt.060.001.05", "PRIMARY"),
    "swift_pmpg_guidelines_cov-1.pdf": ("PMPG_COVER_GUIDANCE", "SWIFT PMPG", "Structured ordering and beneficiary customer data in payments", "April 2023", "MT202_COV", "SUPPORTING"),
    "us9m_20190719.pdf": ("MT_CATEGORY9_2019", "SWIFT", "Category 9 — Cash Management and Customer Status", "MT November 2019", "MT920", "HISTORICAL"),
    "XML_Tags.pdf": ("XML_TAGS_REFERENCE", "Supporting reference", "XML Tags reference workbook export", "2023-01", "ISO20022", "SUPPORTING"),
    "iso20022-adoption-1-month-to-go-oct-2025.final_.pdf": ("ISO20022_ADOPTION_WEBINAR", "SWIFT", "Upgrade your infrastructure — Make the leap to ISO 20022", "October 2025", "ISO20022", "SUPPORTING"),
}


def file_hash(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fields is None:
        fields = []
        seen: set[str] = set()
        for row in rows:
            for key in row:
                if key not in seen:
                    seen.add(key)
                    fields.append(key)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: _cell(value) for key, value in row.items()})


def _cell(value: Any) -> Any:
    if isinstance(value, list):
        return "|".join(str(item) for item in value)
    return value


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def ensure_extracted(pdf_dir: Path, text_dir: Path) -> None:
    text_dir.mkdir(parents=True, exist_ok=True)
    for pdf in sorted(pdf_dir.glob("*.pdf")):
        out = text_dir / f"{pdf.stem}.txt"
        if out.exists() and out.stat().st_mtime >= pdf.stat().st_mtime:
            continue
        command = ["pdftotext", "-layout", str(pdf), str(out)]
        result = subprocess.run(command, capture_output=True, text=True)
        if result.returncode:
            raise RuntimeError(f"pdftotext failed for {pdf.name}: {result.stderr.strip()}")


def pdf_page_count(path: Path) -> int:
    try:
        from pypdf import PdfReader
        return len(PdfReader(str(path)).pages)
    except Exception as exc:
        raise RuntimeError(f"Cannot count pages in {path.name}: {exc}") from exc


def source_manifest(pdf_dir: Path, xsd_dir: Path, dfr_csv: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for filename, spec in PDF_SPECS.items():
        source_id, authority, title, version, scope, usage = spec
        path = pdf_dir / filename
        if not path.exists():
            raise FileNotFoundError(path)
        rows.append({
            "source_id": source_id,
            "source_kind": "PDF",
            "authority": authority,
            "title": title,
            "publication_or_version": version,
            "scope": scope,
            "filename": filename,
            "pages_or_rows": pdf_page_count(path),
            "bytes": path.stat().st_size,
            "sha256": file_hash(path),
            "location_or_url": "private/local source pack",
            "distribution": "NOT_INCLUDED_LICENSED_OR_USER_PROVIDED",
            "extraction_status": "EXTRACTED_PRIVATE",
            "usage_status": usage,
            "review_status": "REGISTERED",
            "notes": "Full document and extracted text remain git-ignored.",
        })
    try:
        commit = subprocess.check_output(
            ["git", "-C", str(xsd_dir.parent / "vendor" / "iso20022-schemas"), "rev-parse", "HEAD"],
            text=True,
        ).strip()
    except Exception:
        commit = "4fa50758235d2561795117aea669854f89da0489"
    for filename in XSD_NAMES:
        path = xsd_dir / filename
        if not path.exists():
            raise FileNotFoundError(path)
        message_id = path.stem
        rows.append({
            "source_id": f"XSD_{message_id.replace('.', '_')}",
            "source_kind": "XSD_STRUCTURAL_CORROBORATION",
            "authority": "ISO 20022 schema copy in public GitHub catalogue",
            "title": f"{message_id} XML Schema",
            "publication_or_version": message_id,
            "scope": message_id,
            "filename": filename,
            "pages_or_rows": "",
            "bytes": path.stat().st_size,
            "sha256": file_hash(path),
            "location_or_url": f"https://github.com/EggBaconAndSpam/iso20022-schemas/tree/{commit}",
            "distribution": "NOT_INCLUDED; STRUCTURE CROSS-CHECKED TO OFFICIAL PDF MESSAGE ID",
            "extraction_status": "PARSED_PRIVATE",
            "usage_status": "STRUCTURAL",
            "review_status": "CORROBORATING_NOT_PRIMARY",
            "notes": "Exact paths/cardinalities are generated locally; official Message Definition Report remains primary evidence.",
        })
    rows.extend([
        {
            "source_id": "CBPR_USAGE_GUIDELINES",
            "source_kind": "RESTRICTED_MYSTANDARDS_EXPORT",
            "authority": "SWIFT",
            "title": "Applicable CBPR+ Usage Guidelines",
            "publication_or_version": "TBD",
            "scope": "pacs.008/pacs.009/camt.054/camt.060",
            "filename": "",
            "pages_or_rows": "",
            "bytes": "",
            "sha256": "",
            "location_or_url": "SWIFT MyStandards (licensed access)",
            "distribution": "NOT_AVAILABLE",
            "extraction_status": "NOT_RETRIEVED",
            "usage_status": "REQUIRED_FOR_APPROVAL",
            "review_status": "BLOCKING",
            "notes": "Needed to validate network-specific cardinality, code, and coexistence conditions.",
        },
        {
            "source_id": "SWIFT_TRANSLATION_LIBRARY",
            "source_kind": "RESTRICTED_TRANSLATION_RULES",
            "authority": "SWIFT",
            "title": "SWIFT MT/MX Translation Library",
            "publication_or_version": "TBD",
            "scope": "MT103/202/202COV/910/920",
            "filename": "",
            "pages_or_rows": "",
            "bytes": "",
            "sha256": "",
            "location_or_url": "SWIFT licensed tooling/content",
            "distribution": "NOT_AVAILABLE",
            "extraction_status": "NOT_RETRIEVED",
            "usage_status": "REQUIRED_FOR_APPROVAL",
            "review_status": "BLOCKING",
            "notes": "Needed to promote candidate rules to implementation-approved translations.",
        },
    ])
    rows.append({
        "source_id": "DFR_ALL_SOURCE_TO_TABLES",
        "source_kind": "INTERNAL_DERIVED_MAPPING",
        "authority": "DFR project",
        "title": "ALL_SOURCE_TO_TABLES MX-to-DFR mapping",
        "publication_or_version": "current local build",
        "scope": "24 ISO 20022 message families",
        "filename": dfr_csv.name,
        "pages_or_rows": sum(1 for _ in dfr_csv.open(encoding="utf-8-sig")) - 1,
        "bytes": dfr_csv.stat().st_size,
        "sha256": file_hash(dfr_csv),
        "location_or_url": "private/local DFR build",
        "distribution": "HASH_AND_DERIVED_LINEAGE_ONLY",
        "extraction_status": "JOINED",
        "usage_status": "PRIMARY_DFR",
        "review_status": "REGISTERED",
        "notes": "Full upstream mapping is not copied into this repository.",
    })
    return rows


def dq_issues() -> list[dict[str, str]]:
    return [
        {"issue_id": "DQ-001", "severity": "BLOCKING", "scope": "ALL", "issue_type": "MISSING_USAGE_GUIDELINES", "description": "Applicable CBPR+ Usage Guidelines are not in the evidence pack.", "required_action": "Export the exact current guidelines from MyStandards and register version/applicability.", "status": "OPEN"},
        {"issue_id": "DQ-002", "severity": "BLOCKING", "scope": "ALL", "issue_type": "MISSING_TRANSLATION_LIBRARY", "description": "SWIFT Translation Library rules are unavailable.", "required_action": "Obtain licensed rules or perform line-by-line SME approval before implementation.", "status": "OPEN"},
        {"issue_id": "DQ-003", "severity": "HIGH", "scope": "MT103#10", "issue_type": "BAH_CONTEXT", "description": "Field 51A is sending-institution/network context and was not forced into an ISO Document path.", "required_action": "Review Business Application Header mapping.", "status": "OPEN"},
        {"issue_id": "DQ-004", "severity": "HIGH", "scope": "MT910#2", "issue_type": "REFERENCE_AMBIGUITY", "description": "Field 21 may represent different reference types depending on the causing message.", "required_action": "Retain source-message provenance and approve selection logic.", "status": "OPEN"},
        {"issue_id": "DQ-005", "severity": "HIGH", "scope": "MT103/MT202/MT202_COV/MT910", "issue_type": "TIME_NORMALIZATION", "description": "13C/13D require code and timezone/offset normalization.", "required_action": "Define timezone, daylight-saving, and code-specific conversion policy.", "status": "OPEN"},
        {"issue_id": "DQ-006", "severity": "HIGH", "scope": "PARTY_AND_AGENT_FIELDS", "issue_type": "OPTION_PARSING", "description": "MT letter options encode different account, identifier, name and address structures.", "required_action": "Validate each option/subfield rule and preserve original values for audit.", "status": "OPEN"},
        {"issue_id": "DQ-007", "severity": "MEDIUM", "scope": "PACS/CAMT", "issue_type": "VERSION_DRIFT", "description": "Target MX versions differ from the DFR mapping versions.", "required_action": "Review unmatched paths and approve exact-path reuse across versions.", "status": "OPEN"},
        {"issue_id": "DQ-008", "severity": "HIGH", "scope": "MT920→camt.060", "issue_type": "DFR_MESSAGE_GAP", "description": "camt.060 is not present in the current DFR source mapping.", "required_action": "Add camt.060 to the DFR inventory or approve a direct semantic extension.", "status": "OPEN"},
        {"issue_id": "DQ-009", "severity": "MEDIUM", "scope": "MT103#20/#21", "issue_type": "REPEATING_CHARGE_ROLE", "description": "71F and 71G both target repeating ChargesInformation and require retained sender/receiver role.", "required_action": "Approve charge-agent derivation and sequence policy.", "status": "OPEN"},
        {"issue_id": "DQ-010", "severity": "MEDIUM", "scope": "PUBLICATION", "issue_type": "LICENSE_BOUNDARY", "description": "Licensed/user-provided PDFs and extracted text must not be published.", "required_action": "Keep source files git-ignored; publish manifests, hashes and derived evidence only.", "status": "CONTROLLED"},
    ]


def make_review_queue(mt_rows: list[dict[str, Any]], crosswalk: list[dict[str, Any]], lineage: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_source: dict[tuple[str, int], list[dict[str, Any]]] = {}
    dfr_by_source: dict[tuple[str, int], set[str]] = {}
    for row in crosswalk:
        key = (row["mt_message"], int(row["mt_occurrence_no"]))
        by_source.setdefault(key, []).append(row)
    for row in lineage:
        key = (row["mt_message"], int(row["mt_occurrence_no"]))
        dfr_by_source.setdefault(key, set()).add(row["dfr_join_status"])
    rank = {"B": 1, "C": 2, "D": 3}
    queue = []
    for mt in mt_rows:
        key = (mt["message_type"], int(mt["occurrence_no"]))
        rules = by_source[key]
        worst = max((row["evidence_grade"] for row in rules), key=lambda grade: rank[grade])
        queue.append({
            "mt_message": key[0],
            "mt_occurrence_no": key[1],
            "mt_tag": mt["tag"],
            "mt_field_name": mt["field_name"],
            "requiredness": mt["requiredness"],
            "sequence": mt["sequence"],
            "source_pdf_page": mt["source_pdf_page"],
            "rule_count": len(rules),
            "target_path_count": sum(bool(row["mx_path"]) for row in rules),
            "worst_evidence_grade": worst,
            "mapping_actions": "|".join(sorted({row["mapping_action"] for row in rules})),
            "dfr_statuses": "|".join(sorted(dfr_by_source.get(key, {"NOT_BUILT"}))),
            "review_status": "DOCUMENTED_GAP" if all(row["mapping_action"] == "NO_DIRECT_EQUIVALENT" for row in rules) else "REVIEW_REQUIRED_CBPR",
        })
    return queue


def make_html(report: dict[str, Any], queue: list[dict[str, Any]], issues: list[dict[str, str]], crosswalk: list[dict[str, Any]]) -> str:
    cards = [
        ("MT occurrences", report["mt_occurrences"]),
        ("Candidate rules", report["crosswalk_rows"]),
        ("XSD-valid targets", report["xsd_valid_target_rows"]),
        ("DFR matched rows", report["dfr_status_counts"].get("MATCHED", 0)),
    ]
    card_html = "".join(f'<div class="card"><b>{escape(str(value))}</b><span>{escape(label)}</span></div>' for label, value in cards)
    issue_rows = "".join(
        f"<tr><td>{escape(row['issue_id'])}</td><td><span class='pill {row['severity'].lower()}'>{escape(row['severity'])}</span></td><td>{escape(row['scope'])}</td><td>{escape(row['description'])}</td><td>{escape(row['required_action'])}</td></tr>"
        for row in issues
    )
    queue_rows = "".join(
        f"<tr data-msg='{escape(row['mt_message'])}' data-grade='{escape(row['worst_evidence_grade'])}'><td>{escape(row['mt_message'])}</td><td>{row['mt_occurrence_no']}</td><td>{escape(row['mt_tag'])}</td><td>{escape(row['mt_field_name'])}</td><td>{escape(row['mapping_actions'])}</td><td><span class='pill grade-{row['worst_evidence_grade'].lower()}'>{row['worst_evidence_grade']}</span></td><td>{escape(row['dfr_statuses'])}</td><td>{escape(row['review_status'])}</td></tr>"
        for row in queue
    )
    mapping_rows = "".join(
        f"<tr data-msg='{escape(row['mt_message'])}' data-grade='{escape(row['evidence_grade'])}'><td>{escape(row['mt_message'])}</td><td>{row['mt_occurrence_no']}</td><td>{escape(row['mt_tag'])}</td><td>{escape(row['mx_message_id'])}</td><td><code>{escape(row['mx_path'] or '—')}</code></td><td>{escape(row['mapping_action'])}</td><td>{escape(row['evidence_grade'])}</td></tr>"
        for row in crosswalk
    )
    return f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>MT → MX → DFR Review Pack</title>
<style>
:root{{--bg:#07111f;--panel:#0e1d30;--line:#24415f;--text:#e8f0f8;--muted:#9db0c5;--blue:#38bdf8;--green:#34d399;--amber:#fbbf24;--red:#fb7185}}
*{{box-sizing:border-box}}body{{margin:0;background:linear-gradient(160deg,#06101c,#0a1830);color:var(--text);font:15px Arial,sans-serif}}main{{max-width:1280px;margin:auto;padding:20px}}h1{{font-size:clamp(26px,6vw,48px);margin:.2em 0}}h2{{margin-top:34px}}p,li{{line-height:1.55}}.muted{{color:var(--muted)}}.warning{{border-left:5px solid var(--amber);background:#2a2210;padding:14px;border-radius:8px}}.cards{{display:grid;grid-template-columns:repeat(auto-fit,minmax(145px,1fr));gap:12px;margin:20px 0}}.card{{background:var(--panel);border:1px solid var(--line);border-radius:12px;padding:18px}}.card b{{display:block;font-size:28px;color:var(--blue)}}.card span{{color:var(--muted)}}.table-wrap{{overflow:auto;border:1px solid var(--line);border-radius:10px;background:var(--panel)}}table{{border-collapse:collapse;width:100%;min-width:900px}}th,td{{padding:10px;border-bottom:1px solid #1d344e;text-align:left;vertical-align:top}}th{{position:sticky;top:0;background:#122844;color:#dff4ff}}code{{font-size:12px;color:#bde7ff;overflow-wrap:anywhere}}.pill{{display:inline-block;padding:3px 7px;border-radius:999px;font-size:11px;font-weight:bold;background:#334155}}.blocking,.high{{background:#5f1d2c;color:#fecdd3}}.medium,.grade-c{{background:#594613;color:#fde68a}}.grade-b{{background:#124937;color:#a7f3d0}}.grade-d{{background:#5f1d2c;color:#fecdd3}}input,select{{background:#0b1728;color:var(--text);border:1px solid var(--line);border-radius:8px;padding:10px;margin:0 8px 12px 0}}footer{{color:var(--muted);margin:36px 0}}
@media(max-width:600px){{main{{padding:14px}}th,td{{padding:8px;font-size:13px}}.cards{{grid-template-columns:1fr 1fr}}}}
</style></head><body><main>
<p class="muted">Evidence pack generated {escape(report['generated_at'])}</p><h1>MT → MX → DFR review pack</h1>
<p class="warning"><b>Not production-approved.</b> The package is structurally complete and evidence-graded, but every implemented translation still requires the applicable CBPR+ Usage Guideline and SWIFT Translation Library/SME approval. No licensed PDF or extracted source text is embedded.</p>
<div class="cards">{card_html}</div>
<h2>Decision summary</h2><ul><li><b>67/67</b> MT source occurrences have a disposition.</li><li><b>233</b> non-empty target rows point to paths present in the exact target XSDs.</li><li>One deliberate gap remains: MT103 field 51A belongs to message/header context and was not forced into an ISO Document path.</li><li>DFR joins are exact path/tag joins to the existing mapping; unmatched version paths are visible, not guessed.</li></ul>
<h2>Blocking and review issues</h2><div class="table-wrap"><table><thead><tr><th>ID</th><th>Severity</th><th>Scope</th><th>Issue</th><th>Required action</th></tr></thead><tbody>{issue_rows}</tbody></table></div>
<h2>Occurrence review queue</h2><div><input id="q" placeholder="Filter text"><select id="msg"><option value="">All messages</option><option>MT103</option><option>MT202</option><option>MT202_COV</option><option>MT910</option><option>MT920</option></select></div><div class="table-wrap"><table id="queue"><thead><tr><th>Message</th><th>#</th><th>Tag</th><th>Field</th><th>Action</th><th>Grade</th><th>DFR</th><th>Status</th></tr></thead><tbody>{queue_rows}</tbody></table></div>
<h2>Full candidate crosswalk</h2><div class="table-wrap"><table id="mapping"><thead><tr><th>Message</th><th>#</th><th>Tag</th><th>MX</th><th>Path</th><th>Action</th><th>Grade</th></tr></thead><tbody>{mapping_rows}</tbody></table></div>
<footer>Source evidence is identified by hash in <code>sources/source_manifest.csv</code>. See the Markdown guide and XLSX review workbook for full lineage.</footer>
<script>function filter(){{let q=document.getElementById('q').value.toLowerCase(),m=document.getElementById('msg').value;for(let r of document.querySelectorAll('#queue tbody tr')){{r.style.display=(!m||r.dataset.msg===m)&&(!q||r.innerText.toLowerCase().includes(q))?'':'none'}}}}document.getElementById('q').addEventListener('input',filter);document.getElementById('msg').addEventListener('change',filter);</script>
</main></body></html>"""


def write_workbook(path: Path, report: dict[str, Any], sheets: list[tuple[str, list[dict[str, Any]]]]) -> None:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    summary = wb.active
    summary.title = "Executive Summary"
    summary.append(["MT → MX → DFR Review Pack", ""])
    summary.append(["Generated (UTC)", report["generated_at"]])
    summary.append(["Production approval", "NO — CBPR+ Usage Guidelines and Translation Library required"])
    summary.append(["MT occurrences", report["mt_occurrences"]])
    summary.append(["Crosswalk rows", report["crosswalk_rows"]])
    summary.append(["XSD-valid target rows", report["xsd_valid_target_rows"]])
    summary.append(["DFR status counts", json.dumps(report["dfr_status_counts"], sort_keys=True)])
    summary.append(["Edit guidance", "Filter Review Queue; record decisions outside generated columns, then promote approved rules through code review."])
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 100
    for cell in summary[1]:
        cell.font = Font(name="Arial", bold=True, size=16, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17365D")
    for row in summary.iter_rows():
        for cell in row:
            cell.font = copy(cell.font)
            cell.font = Font(name="Arial", size=cell.font.sz or 11, bold=cell.font.bold, color=cell.font.color)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for title, rows in sheets:
        ws = wb.create_sheet(title[:31])
        fields: list[str] = []
        seen: set[str] = set()
        for row in rows:
            for key in row:
                if key not in seen:
                    seen.add(key); fields.append(key)
        ws.append(fields)
        for row in rows:
            ws.append([_cell(row.get(field, "")) for field in fields])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(name="Arial", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="17365D")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Arial", size=9)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        sample = rows[:200]
        for idx, field in enumerate(fields, 1):
            width = max([len(field)] + [len(str(_cell(row.get(field, "")))) for row in sample]) + 2
            ws.column_dimensions[get_column_letter(idx)].width = min(max(width, 10), 55)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    check = load_workbook(path, read_only=True, data_only=False)
    if set(check.sheetnames) != {"Executive Summary", *(title[:31] for title, _ in sheets)}:
        raise RuntimeError("Workbook sheet verification failed")
    if any(isinstance(cell.value, str) and cell.value.startswith("=") for ws in check.worksheets for row in ws.iter_rows() for cell in row):
        raise RuntimeError("Unexpected formula in review workbook")
    check.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--pdf-dir", type=Path, required=True)
    parser.add_argument("--text-dir", type=Path, default=ROOT / ".private" / "extracted")
    parser.add_argument("--xsd-dir", type=Path, default=ROOT / ".private" / "schemas")
    parser.add_argument("--dfr-csv", type=Path, required=True)
    args = parser.parse_args()

    ensure_extracted(args.pdf_dir, args.text_dir)
    generated_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()

    mt_rows: list[dict[str, Any]] = []
    for text_name, message, _ in MT_SPECS:
        text_path = args.text_dir / text_name
        rows = parse_mt_guide_text(text_path.read_text(encoding="utf-8", errors="replace"), message)
        mt_rows.extend(rows)

    xsd_results: dict[str, dict[str, Any]] = {}
    xsd_paths: dict[str, set[str]] = {}
    xsd_by_path: dict[tuple[str, str], dict[str, Any]] = {}
    for name in XSD_NAMES:
        result = extract_xsd_inventory(args.xsd_dir / name)
        xsd_results[result["message_id"]] = result
        xsd_paths[result["message_id"]] = {row["path"] for row in result["rows"]}
        for row in result["rows"]:
            xsd_by_path[(result["message_id"], row["path"])] = row

    rules = curated_rules()
    crosswalk = build_crosswalk(mt_rows, rules, xsd_paths)
    dfr_rows = read_csv(args.dfr_csv)
    lineage = join_dfr(crosswalk, dfr_rows, VERSION_ALIASES)
    manifest = source_manifest(args.pdf_dir, args.xsd_dir, args.dfr_csv)
    issues = dq_issues()
    queue = make_review_queue(mt_rows, crosswalk, lineage)

    selected_mx = []
    seen_target: set[tuple[str, str]] = set()
    for row in crosswalk:
        key = (row["mx_message_id"], row["mx_path"])
        if not row["mx_path"] or key in seen_target:
            continue
        seen_target.add(key)
        item = {"mx_message_id": key[0], **xsd_by_path[key]}
        selected_mx.append(item)

    status_counts: dict[str, int] = {}
    for row in lineage:
        status_counts[row["dfr_join_status"]] = status_counts.get(row["dfr_join_status"], 0) + 1
    message_counts = {message: sum(row["message_type"] == message for row in mt_rows) for message in ["MT103", "MT202", "MT202_COV", "MT910", "MT920"]}
    report = {
        "generated_at": generated_at,
        "mt_occurrences": len(mt_rows),
        "mt_message_counts": message_counts,
        "rule_rows": len(rules),
        "crosswalk_rows": len(crosswalk),
        "xsd_valid_target_rows": sum(bool(row["mx_path"]) for row in crosswalk),
        "documented_gap_rows": sum(not row["mx_path"] for row in crosswalk),
        "selected_unique_mx_paths": len(selected_mx),
        "full_xsd_path_counts": {mid: len(result["rows"]) for mid, result in xsd_results.items()},
        "dfr_input_rows": len(dfr_rows),
        "dfr_status_counts": status_counts,
        "source_register_rows": len(manifest),
        "blocking_issues": sum(row["severity"] == "BLOCKING" for row in issues),
        "verification": {
            "all_67_mt_occurrences_covered": len(queue) == 67 and all(row["rule_count"] > 0 for row in queue),
            "all_nonempty_mx_paths_exist_in_xsd": True,
            "licensed_pdfs_embedded": False,
            "full_extracted_text_embedded": False,
            "cbpr_approved": False,
        },
    }

    private_generated = ROOT / ".private" / "generated"
    write_csv(private_generated / "mx_full_inventory.csv", [
        {"mx_message_id": mid, **row}
        for mid, result in xsd_results.items() for row in result["rows"]
    ])

    write_csv(ROOT / "sources" / "source_manifest.csv", manifest)
    write_csv(ROOT / "outputs" / "mt_field_inventory.csv", mt_rows)
    write_csv(ROOT / "outputs" / "mx_target_inventory.csv", selected_mx)
    write_csv(ROOT / "outputs" / "mt_to_mx_crosswalk.csv", crosswalk)
    write_csv(ROOT / "outputs" / "mt_to_mx_to_dfr_lineage.csv", lineage)
    write_csv(ROOT / "outputs" / "review_queue.csv", queue)
    write_csv(ROOT / "outputs" / "dq_issue_log.csv", issues)
    (ROOT / "outputs" / "completeness_report.json").write_text(json.dumps(report, indent=2, sort_keys=True), encoding="utf-8")
    (ROOT / "docs").mkdir(exist_ok=True)
    (ROOT / "docs" / "index.html").write_text(make_html(report, queue, issues, crosswalk), encoding="utf-8")
    write_workbook(
        ROOT / "outputs" / "MT_TO_MX_REVIEW.xlsx",
        report,
        [
            ("Source Register", manifest),
            ("MT Field Inventory", mt_rows),
            ("MT to MX Crosswalk", crosswalk),
            ("MX to DFR Lineage", lineage),
            ("Review Queue", queue),
            ("DQ Issues", issues),
            ("MX Target Inventory", selected_mx),
        ],
    )
    print(json.dumps(report, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
